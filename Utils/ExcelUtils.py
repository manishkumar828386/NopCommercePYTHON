from multiprocessing.pool import worker

import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getCellCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName, rownum,columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row = rownum, column = columnum).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row = rownum , column = columnno).value = data
    workbook.save(file)